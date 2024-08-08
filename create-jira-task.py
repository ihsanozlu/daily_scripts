from jira import JIRA
from openpyxl import load_workbook
from datetime import datetime
from workdays import workday

# Priority mapping dictionary
priority_mapping = {
    1: "Important",
    2: "Cosmetic",
    3: "Nice to Have"
}

# Load Excel file
workbook = load_workbook('/Users/ihsan/Documents/data-projects.xlsx')
sheet = workbook.active

# Connect to Jira
options = {
    'server': 'https://yourserver.atlassian.net'
}

try:
    jira = JIRA(options, basic_auth=('your@email.net', 'yourapitoken'))
    jira.projects()  # Test connection by fetching projects
except Exception as e:
    print("Failed to connect to Jira:", e)
    exit()

# Iterate through rows in Excel
for row in sheet.iter_rows(min_row=2, values_only=True):
    TASKS, COMPLEXITY, PRIORITY, START, END, WORKDAYS, PROJECT_KEY, PROJECT, SUBPROJECT = row[:9]

    # Declare start_date variable
    start_date = None

    # Check if start date is provided
    if START is None:
        print("Start date is not provided. Creating Epic without due date.")
        end_date = None
    else:
        # Convert start date to datetime object
        if isinstance(START, str):
            try:
                start_date = datetime.strptime(START, "%d.%m.%Y")
            except ValueError:
                print(f"Invalid start date format: '{START}'")
                continue
        else:
            print(f"Invalid start date format: '{START}'")
            continue

        # Calculate end date if not provided
        if END is None:
            if WORKDAYS is not None:
                try:
                    end_date = workday(start_date, WORKDAYS)
                except Exception as e:
                    print("Failed to calculate end date:", e)
                    continue
            else:
                print("Workdays is not provided. Skipping row.")
                continue
        else:
            print("End date is provided. Skipping row.")
            continue

    # Map priority to Jira priority
    jira_priority = priority_mapping.get(PRIORITY)
    if jira_priority is None:
        print(f"Invalid priority: '{PRIORITY}'. Skipping row.")
        continue

    # Create Epic summary
    epic_summary = f"{PROJECT} - {SUBPROJECT} | {TASKS}"

    # Create Epic
    epic_dict = {
        'project': {'key': PROJECT_KEY},
        'summary': epic_summary,
        'description': f"Complexity: {COMPLEXITY}\nPriority: {jira_priority}\nStart Date: {start_date.strftime('%Y-%m-%d')}" if start_date else f"Complexity: {COMPLEXITY}\nPriority: {jira_priority}",
        'issuetype': {'name': 'Epic'},
        'priority': {'name': jira_priority}  # Set the priority of the epic
    }

    # Include due date if end date is provided
    if end_date:
        epic_dict['duedate'] = end_date.strftime('%Y-%m-%d')

    try:
        epic = jira.create_issue(fields=epic_dict)
        print('Created epic:', epic.key)
    except Exception as e:
        print("Failed to create epic:", e)

